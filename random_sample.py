
### 随机抽取指定数量的医生和患者，用于empirical.py训练


import openpyxl
import random


# 定义一个函数，获取excel指定行的数据。
from openpyxl import load_workbook


def get_row_value(ws, row):
    col_num = ws.max_column
    row_data = []
    for i in range(1, col_num + 1):
        cell_value = ws.cell(row=row, column=i).value
        row_data.append(cell_value)
    return row_data

#随机读取指定数量的患者
# file_name_list = ['./data/final_data_time636.xlsx']  # 将文件名集合放在一个列表。

#随机读取指定数量的医生
file_name_list = ['./data/vas_list_answers5.xlsx']
for file_name in file_name_list:

    # 读取
    wb = load_workbook(file_name)
    sheet = wb.active
    row_num = sheet.max_row
    random_num = random.sample(range(2, row_num + 1), 20)  # 随机抽取20个样本，第一行是表头，不取。

    # 写入一个新的excel表格
    wb2 = openpyxl.Workbook()
    sheet2 = wb.active
    sheet2.append(get_row_value(sheet, 1))  # 自定义的函数，传入两个参数，一个是前文读取的文件，一个是要读取的行数。
    # 读取随机抽取的行数并写入。
    for i in random_num:
        row = get_row_value(sheet2, i)
        sheet.append(row)
    sheet.append(['生成的随机数为：'] + random_num)

    # 保存
    # out_file_name = file_name + '_out.xlsx'
    # out_file_name = './data/random_patients/10_patients.xlsx'

    out_file_name = './data/random_doctors/20_doctors.xlsx'

    wb.save(out_file_name)
    print('抽样成功')
